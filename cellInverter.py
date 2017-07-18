import openpyxl
import sys
import os

if len(sys.argv) < 2:
    print(r'''Usage:
        python cellInverter.py [file]
        python cellInverter.py [file] [output]
            ''')
    exit()
print('Opening file')
wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.get_active_sheet()

output = openpyxl.Workbook()
outputsheet = output.get_active_sheet()
print('Inverting cells')
for x in range(1, sheet.max_row+1):
    for y in range(1, sheet.max_column + 1):
        outputsheet.cell(row=y, column=x).value = sheet.cell(row=x, column=y).value

if len(sys.argv) < 3:
    output.save('Inverted' + os.path.basename(sys.argv[1]))
    print('File saved as Inverted' + os.path.basename(sys.argv[1]))
else:
    output.save(sys.argv[2])
    print('File saved as ' + sys.argv[2])

print('Done!')

